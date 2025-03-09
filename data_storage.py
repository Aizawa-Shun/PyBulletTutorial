"""
データ保存関連機能
--------------
このモジュールは、シミュレーションデータの保存を担当します。
"""

import os
import openpyxl

def save_data_to_excel(joint_angles_data):
    """関節角度データをExcelファイルに保存する"""
    # ディレクトリが存在しない場合は作成
    os.makedirs("./sheet", exist_ok=True)
    
    # ワークブックの作成またはロード
    excel_path = "./sheet/data.xlsx"
    if not os.path.isfile(excel_path):
        wb = openpyxl.Workbook()
        wb.save(excel_path)
    
    # ワークブックをロードしシートを作成
    wb = openpyxl.load_workbook(excel_path)
    
    # すでに存在する場合はシートを削除して重複を避ける
    if "joint_angles" in wb.sheetnames:
        wb.remove(wb["joint_angles"])
    
    # 新しいシートを作成
    ws = wb.create_sheet(title="joint_angles")
    
    # データの書き込み
    for y, row in enumerate(joint_angles_data):
        for x, value in enumerate(row):
            ws.cell(y+1, x+1).value = value
    
    wb.save(excel_path)
    print(f"データを {excel_path} に保存しました")
